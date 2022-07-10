from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook


def _get_sheet_copy(path: str, sheet_name: str, version: int = 0) -> Worksheet:
    source_workbook = load_workbook(path, data_only=True)
    if not version:
        source_workbook.save("workbook_copy.xlsx")
        return load_workbook("workbook_copy.xlsx")[sheet_name]

    source_workbook.save("workbook_copy_" + str(version) + ".xlsx")
    return load_workbook("workbook_copy_" + str(version) + ".xlsx")[sheet_name]


def get_sample_distributor_description_as_id_sheet(version: int = 0) -> Worksheet:
    return _get_sheet_copy(
        "/Users/stanleykurniawan/Downloads/BSD MEI 2022.xlsx", "BSD MEI 2022", version
    )


def get_sample_distributor_code_as_id_sheet(version: int = 0) -> Worksheet:
    return _get_sheet_copy(
        "/Users/stanleykurniawan/Downloads/saputra alsut mei 2022.xlsx",
        "Sheet1",
        version,
    )


def get_sample_master_sheet(version: int = 0) -> Worksheet:
    return _get_sheet_copy(
        "/Users/stanleykurniawan/Downloads/MASTER NY JANUARI 2022.xlsx",
        "Upload",
        version,
    )
